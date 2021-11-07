
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from check_xlsx import cleaning_dir_xlsx
from main import DIR_XLSX

cleaning_dir_xlsx()

filename_xlsx = f'СводДоходов_Ульяновск_10.xlsx'

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Свод Доходов'

ws['A1'] = 'Свод доходов по Ульяновскому региону'
ws['A1'].style = 'Title'
# ws['A2'] = 'В разрезе подразделений'
ws['B3'] = 'Период'
ws['B3'].style = 'Headline 4'
ws['C3'].style = 'Headline 4'
ws['D3'].style = 'Headline 4'

ws['A5'] = 'Подразделение'
ws['A6'] = 'Вид услуги'
ws['B5'] = 'Перевезенная сумма'
ws['C5'] = 'Доход'
ws['D5'] = 'Доход с НДС'
ws.merge_cells('B5:B6')
ws.merge_cells('C5:C6')
ws.merge_cells('D5:D6')
ws['A5'].style = 'Accent1'
ws['A6'].style = 'Accent1'
ws['A5'].style = 'Accent1'
ws['B5'].style = 'Accent1'
ws['C5'].style = 'Accent1'
ws['D5'].style = 'Accent1'

ws['B5'].alignment = Alignment(horizontal="left", vertical="center")
ws['C5'].alignment = Alignment(horizontal="left", vertical="center")
ws['D5'].alignment = Alignment(horizontal="left", vertical="center")

ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20


# for y in tuple_:
#     print(y)
#     for i in range(len(y)):
#         print(i)
#         ws.append(y[i])
#
# ws.append(sum_tuple)


print(f"Запись файла '{filename_xlsx}'")
wb.save(DIR_XLSX + '//' + filename_xlsx)
